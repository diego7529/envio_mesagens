import pandas as pd
import pywhatkit as kit
import pyautogui
import time
from openpyxl import load_workbook
import os
from datetime import datetime

def horario_comercial():
    current_time = datetime.now().time()
    start_time = current_time.replace(hour=8, minute=0, second=0, microsecond=0)
    end_time = current_time.replace(hour=18, minute=0, second=0, microsecond=0)
    return start_time <= current_time <= end_time

contacts = pd.read_excel('numeros.xlsx')
mensagem_inicial = ""

#video_path = "C:\\Projects\\python\\envios_mensagem\\video_teste.mp4"

workbook = load_workbook('numeros.xlsx')
sheet = workbook.active

count = 0

for index, row in contacts.iterrows():  
    while not horario_comercial():
        print("Fora do horário. Modo de espera...")
        time.sleep(60)  

    numero = row['Número']
    mensagem = row['Mensagem']
    video_path = row['Video']
    verificacao = row['Enviado']

    if pd.isnull(numero):
        sheet.cell(row=index + 2, column=4).value = "NÃO"
        workbook.save('numeros.xlsx')
        continue

    if pd.notnull(verificacao):
        print("Proxima linha!")
        continue

    kit.sendwhatmsg_instantly(numero, mensagem_inicial)

    pyautogui.click(x=764, y=956)  
    time.sleep(1)
    pyautogui.click(x=839, y=688) 
    time.sleep(1)
    pyautogui.write(video_path)
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.click(x=884, y=960)  
    time.sleep(1)
    pyautogui.typewrite(mensagem, interval=0.05)
    pyautogui.press('enter')
    pyautogui.hotkey('ctrl', 'w')
    pyautogui.press('enter')
    sheet.cell(row=index + 2, column=4).value = "SIM"
    workbook.save('numeros.xlsx')
    count = count + 1

    time.sleep(5)

print(f"Mensagens enviadas com sucesso : {count}")
